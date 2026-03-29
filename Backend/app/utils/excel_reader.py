
from openpyxl import load_workbook
from io import BytesIO
from openpyxl import load_workbook
from io import BytesIO
from datetime import datetime

def extract_emails(upload_file):
    """
    Extract user data from excel file including name, email, contact, and role-specific fields.
    Returns a list of dicts with all available user information.
    """
    file_bytes = upload_file.read()
    excel_file = BytesIO(file_bytes)

    wb = load_workbook(excel_file)
    sheet = wb.active

    users = []
    header = [cell.value for cell in sheet[1]]

    if "email" not in header:
        raise ValueError("Excel must contain 'email' column")

    # Map column indices
    email_col = header.index("email")
    personal_email_col = header.index("personal_email") if "personal_email" in header else None
    first_name_col = header.index("first_name") if "first_name" in header else None
    last_name_col = header.index("last_name") if "last_name" in header else None
    mobile_no_col = header.index("mobile_no") if "mobile_no" in header else None
    qualification_col = header.index("qualification") if "qualification" in header else None
    years_exp_col = header.index("years_of_experience") if "years_of_experience" in header else None
    subject_code_col = header.index("subject_code") if "subject_code" in header else None
    subject_name_col = header.index("subject_name") if "subject_name" in header else None
    branch_col = header.index("branch") if "branch" in header else None
    address_col = header.index("address") if "address" in header else None

    email_counts = {}

    for row in sheet.iter_rows(min_row=2):
        email = row[email_col].value
        
        if not email:
            continue
        
        email = str(email).strip().lower()
        if not email:
            continue

        email_counts[email] = email_counts.get(email, 0) + 1

        user_data = {"email": email}
        
        # Extract optional fields
        if personal_email_col is not None and row[personal_email_col].value:
            user_data["personal_email"] = str(row[personal_email_col].value).strip()
        if first_name_col is not None and row[first_name_col].value:
            user_data["first_name"] = str(row[first_name_col].value).strip()
        if last_name_col is not None and row[last_name_col].value:
            user_data["last_name"] = str(row[last_name_col].value).strip()
        if mobile_no_col is not None and row[mobile_no_col].value:
            user_data["mobile_no"] = str(row[mobile_no_col].value).strip()
        if qualification_col is not None and row[qualification_col].value:
            user_data["qualification"] = str(row[qualification_col].value).strip()
        if years_exp_col is not None and row[years_exp_col].value:
            user_data["years_of_experience"] = int(row[years_exp_col].value) if isinstance(row[years_exp_col].value, (int, float)) else None
        if subject_code_col is not None and row[subject_code_col].value:
            user_data["subject_code"] = str(row[subject_code_col].value).strip()
        if subject_name_col is not None and row[subject_name_col].value:
            user_data["subject_name"] = str(row[subject_name_col].value).strip()
        if branch_col is not None and row[branch_col].value:
            user_data["branch"] = str(row[branch_col].value).strip()
        if address_col is not None and row[address_col].value:
            user_data["address"] = str(row[address_col].value).strip()

        users.append(user_data)

    duplicates = [e for e, c in email_counts.items() if c > 1]
    return {"users": users, "duplicates": duplicates}


def extract_academic_rows(upload_file):
    file_bytes = upload_file.file.read()
    excel = BytesIO(file_bytes)

    wb = load_workbook(excel)
    sheet = wb.active

    header = [cell.value for cell in sheet[1]]
    rows = []

    for row in sheet.iter_rows(min_row=2):
        data = dict(zip(header, [cell.value for cell in row]))

        rows.append({
            "srno": str(data["srno"]),
            "branch": data["branch"],
            "batch": data["batch"],
            "course": data["course"],
            "year": int(data["year"]),
            "semester": int(data["semester"]),
            "section": data["section"],
            "type": data["type"],
            "quota": data.get("quota"),
            "admission_date": data["admission_date"],
            "status": data["status"]
        })

    return rows
