from openpyxl import load_workbook
from io import BytesIO

from app.models.internal_marks import InternalMarks
from app.models.external_marks import ExternalMarks
from app.models.student import Student
from app.models.academic import Academic
from app.models.semester_result import SemesterResult
from app.utils.marks_calculator import (
    calc_mid_total,
    calc_final_internal,
    grade_from_score,
    calc_sgpa,
    round2,
)
from app.utils.academic_year import resolve_year

def upload_internal_marks_excel(
    db,
    file,
    year: int,
    semester: int,
    subject_code: str,
    faculty_email: str
):
    file.file.seek(0)
    wb = load_workbook(BytesIO(file.file.read()))
    sheet = wb.active

    for row in sheet.iter_rows(min_row=2, values_only=True):
        (
            roll_no,
            subject_code,
            subject_name,
            openbook1,
            openbook2,
            objective1,
            objective2,
            descriptive1,
            descriptive2,
            seminar1,
            seminar2
        ) = row
        if not roll_no or not subject_code:
            continue
        subject_code = str(subject_code).strip()
        subject_name = str(subject_name).strip() if subject_name is not None else subject_code

        student = db.query(Student).filter(
            Student.roll_no == str(roll_no)
        ).first()
        if not student:
            continue

        record = db.query(InternalMarks).filter(
            InternalMarks.sid == student.id,
            InternalMarks.subject_code == subject_code,
            InternalMarks.year == year,
            InternalMarks.semester == semester
        ).first()

        mid1 = calc_mid_total(openbook1, objective1, descriptive1, seminar1)
        mid2 = calc_mid_total(openbook2, objective2, descriptive2, seminar2)

        if record:
            record.subject_name = subject_name
            record.openbook1 = openbook1
            record.openbook2 = openbook2
            record.objective1 = objective1
            record.objective2 = objective2
            record.descriptive1 = descriptive1
            record.descriptive2 = descriptive2
            record.seminar1 = seminar1
            record.seminar2 = seminar2
            record.mid1 = mid1
            record.mid2 = mid2
        else:
            db.add(InternalMarks(
                sid=student.id,
                srno=student.roll_no,
                subject_code=subject_code,
                subject_name=subject_name,
                year=year,
                semester=semester,
                openbook1=openbook1,
                openbook2=openbook2,
                objective1=objective1,
                objective2=objective2,
                descriptive1=descriptive1,
                descriptive2=descriptive2,
                seminar1=seminar1,
                seminar2=seminar2,
                mid1=mid1,
                mid2=mid2,
                entered_by=faculty_email
            ))

    db.commit()
def upload_external_marks_excel(db, file, batch: str, semester: int, admin_email: str, branch: str = None, section: str = None):
    """
    Upload external marks from Excel file.
    
    Parameters:
    - batch: Academic batch (e.g., "2022-26")
    - semester: Semester number (1-8)
    - branch: Branch code (e.g., "CSE", "CSM") - can be None for all branches
    - section: Section (A, B, C) - can be None for all sections
    
    Expected Excel columns: rollno, subjectcode, subjectname, credits, externalmarks
    """
    file.file.seek(0)
    wb = load_workbook(BytesIO(file.file.read()))
    sheet = wb.active

    processed_students = {}  # To track SGPA/CGPA calculation per student
    processed_records_count = 0
    
    for row in sheet.iter_rows(min_row=2, values_only=True):
        try:
            roll_no, subject_code, subject_name, credits, external_marks = row
            if not roll_no or not subject_code:
                continue
                
            # Normalize inputs
            roll_no = str(roll_no).strip()
            subject_code = str(subject_code).strip()
            subject_name = str(subject_name).strip() if subject_name else subject_code
            
            try:
                credits = int(credits) if credits else 0
            except (ValueError, TypeError):
                credits = 0
                
            try:
                external_val = int(float(external_marks)) if external_marks else 0
            except (ValueError, TypeError):
                external_val = 0

            # Find student
            student = db.query(Student).filter(
                Student.roll_no == roll_no
            ).first()
            if not student:
                continue

            # Find academic record with batch, branch, and section filters
            academic_query = db.query(Academic).filter(
                Academic.sid == student.id,
                Academic.batch == batch
            )
            
            if branch and branch != "All":
                academic_query = academic_query.filter(Academic.branch == branch)
            
            if section and section != "All":
                academic_query = academic_query.filter(Academic.section == section)
                
            academic = academic_query.first()
            
            if not academic:
                continue

            year = academic.year
            student_branch = academic.branch
            student_section = academic.section

            # Calculate grade based on external marks (out of 70)
            grade, _ = grade_from_score(external_val)

            # Create or update external marks record
            record = db.query(ExternalMarks).filter(
                ExternalMarks.sid == student.id,
                ExternalMarks.subject_code == subject_code,
                ExternalMarks.semester == semester,
                ExternalMarks.batch == batch
            ).first()

            if record:
                # Update existing record
                record.subject_name = subject_name
                record.credits = credits
                record.external_marks = external_val
                record.grade = grade
                record.entered_by = admin_email
            else:
                # Create new record
                db.add(ExternalMarks(
                    sid=student.id,
                    srno=roll_no,  # For backward compatibility
                    rollno=roll_no,
                    subject_code=subject_code,
                    subject_name=subject_name,
                    batch=batch,
                    branch=student_branch,
                    section=student_section,
                    year=year,
                    semester=semester,
                    credits=credits,
                    external_marks=external_val,
                    grade=grade,
                    entered_by=admin_email
                ))
            
            processed_records_count += 1

            # Track for SGPA calculation
            student_key = student.id
            if student_key not in processed_students:
                processed_students[student_key] = {
                    "total_points": 0.0,
                    "total_credits": 0,
                    "has_fail": False,
                    "branch": student_branch,
                    "section": student_section,
                    "year": year
                }
            
            # Convert external marks to GPA scale (assuming direct mapping or 10-point scale)
            gpa_value = (external_val / 70.0) * 10.0  # Convert 70-point scale to 10-point GPA
            processed_students[student_key]["total_points"] += gpa_value * credits
            processed_students[student_key]["total_credits"] += credits
            
            if grade == "F":
                processed_students[student_key]["has_fail"] = True
                
        except Exception as e:
            print(f"Error processing row: {e}")
            continue

    db.commit()

    # Calculate and update semester results
    for student_id, data in processed_students.items():
        if data["total_credits"] > 0:
            sgpa = calc_sgpa(data["total_points"], data["total_credits"])
        else:
            sgpa = 0.0

        result_status = "FAIL" if data["has_fail"] else "PASS"

        # Check for existing semester result
        existing = db.query(SemesterResult).filter(
            SemesterResult.sid == student_id,
            SemesterResult.semester == semester,
            SemesterResult.batch == batch
        ).first()

        if existing:
            # Update existing record
            existing.sgpa = sgpa
            existing.total_credits = data["total_credits"]
            existing.result_status = result_status
        else:
            # Create new semester result
            student_record = db.query(Student).filter(Student.id == student_id).first()
            db.add(SemesterResult(
                sid=student_id,
                srno=student_record.roll_no if student_record else "",  # For backward compatibility
                rollno=student_record.roll_no if student_record else "",
                batch=batch,
                branch=data["branch"],
                section=data["section"],
                year=data["year"],
                semester=semester,
                sgpa=sgpa,
                cgpa=sgpa,  # Initially set CGPA = SGPA (should be calculated from all semesters)
                total_credits=data["total_credits"],
                result_status=result_status
            ))

    db.commit()



def calculate_sgpa_and_promote(db, student_id: int, semester: int):
    pass

def promote_student_after_external(db, academic: Academic):
    pass

