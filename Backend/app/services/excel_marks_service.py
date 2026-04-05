"""
Enhanced marks service with proper internal and external marks calculations
following institutional grading system.
"""

from typing import Any, Iterator, Optional, Tuple

from openpyxl import load_workbook
from io import BytesIO
from app.models.internal_marks import InternalMarks
from app.models.external_marks import ExternalMarks
from app.models.student import Student
from app.models.academic import Academic
from app.models.course_grade import CourseGrade, SemesterGrade
from app.utils.marks_calculator import (
    calculate_mid_marks,
    calculate_final_internal_marks,
    calculate_semester_marks,
    grade_from_percentage,
    calc_sgpa,
    cgpa_to_percentage,
    round2,
)
from app.utils.academic_year import resolve_year


def _norm_header_cell(cell: Any) -> str:
    if cell is None:
        return ""
    return str(cell).strip().lower().replace("_", "").replace(" ", "")


def _canonical_excel_col(norm: str) -> Optional[str]:
    if norm in ("rollno", "rollnumber", "srno", "roll"):
        return "rollno"
    if norm in ("subjectcode", "subcode", "subject_code"):
        return "subjectcode"
    if norm in ("subjectname", "subject", "subject_name", "name", "subname"):
        return "subjectname"
    if norm in ("credits", "credit"):
        return "credits"
    if norm in (
        "externalmarks",
        "external",
        "extmarks",
        "externalmark",
        "semesterexam",
        "semmarks",
        "theory",
    ):
        return "externalmarks"
    return None


def iter_external_marks_sheet_rows(sheet) -> Iterator[Tuple[Any, Any, Any, Any, Any]]:
    """
    Yield (roll_no, subject_code, subject_name, credits, external_marks) per row.
    Uses the first row as headers when rollno + subjectcode are detected; otherwise
    assumes fixed order: rollno, subjectcode, subjectname, credits, externalmarks.
    """
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return
    header_norm = [_norm_header_cell(c) for c in rows[0]]
    colmap = {}
    for i, h in enumerate(header_norm):
        key = _canonical_excel_col(h)
        if key and key not in colmap:
            colmap[key] = i
    if "rollno" in colmap and "subjectcode" in colmap:
        for row in rows[1:]:
            if not row or all(x is None for x in row):
                continue

            def get_col(name: str):
                j = colmap.get(name)
                if j is None or j >= len(row):
                    return None
                return row[j]

            yield (
                get_col("rollno"),
                get_col("subjectcode"),
                get_col("subjectname"),
                get_col("credits"),
                get_col("externalmarks"),
            )
        return
    for row in rows[1:]:
        if not row or all(x is None for x in row):
            continue
        cells = list(row) + [None] * 5
        yield cells[0], cells[1], cells[2], cells[3], cells[4]


def convert_mark_value(value):
    """
    Convert mark value from Excel. 
    Handles:
    - "AB" (Absent) -> returns 0 and True flag for absent
    - Numbers -> returns converted int
    - Returns (value, is_absent)
    """
    if not value:
        return 0, False
    
    value_str = str(value).strip().upper()
    
    if value_str == "AB":
        return 0, True
    
    try:
        return int(float(value)), False
    except (ValueError, TypeError):
        return 0, True


def upload_internal_marks_excel(
    db,
    file,
    year: int,
    semester: int,
    subject_code: str,
    faculty_email: str
):
    """
    Faculty uploads internal marks with components.
    
    Excel columns: rollno, subjectcode, subjectname, 
    objective1, descriptive1, openbook1, seminar1,
    objective2, descriptive2, openbook2, seminar2
    
    Calculates: mid1, mid2, final_internal_marks automatically
    """
    file.file.seek(0)
    wb = load_workbook(BytesIO(file.file.read()))
    sheet = wb.active

    processed_count = 0
    
    for row in sheet.iter_rows(min_row=2, values_only=True):
        try:
            # Unpack row
            (
                roll_no,
                subject_code,
                subject_name,
                objective1,
                descriptive1,
                openbook1,
                seminar1,
                objective2,
                descriptive2,
                openbook2,
                seminar2
            ) = row
            
            if not roll_no or not subject_code:
                continue
                
            roll_no = str(roll_no).strip()
            subject_code = str(subject_code).strip()
            subject_name = str(subject_name).strip() if subject_name else subject_code

            # Find student
            student = db.query(Student).filter(
                Student.roll_no == roll_no
            ).first()
            
            if not student:
                continue

            # Convert marks, handling AB (Absent) marking
            obj1_val, is_absent1 = convert_mark_value(objective1)
            desc1_val, _ = convert_mark_value(descriptive1)
            open1_val, _ = convert_mark_value(openbook1)
            sem1_val, _ = convert_mark_value(seminar1)
            
            obj2_val, is_absent2 = convert_mark_value(objective2)
            desc2_val, _ = convert_mark_value(descriptive2)
            open2_val, _ = convert_mark_value(openbook2)
            sem2_val, _ = convert_mark_value(seminar2)
            
            # If either mid is marked absent, mid = 0
            is_mid_absent = is_absent1 or is_absent2
            
            # Calculate mid1 from components
            mid1 = calculate_mid_marks(obj1_val, desc1_val, open1_val, sem1_val) if not is_absent1 else 0
            
            # Calculate mid2 from components
            mid2 = calculate_mid_marks(obj2_val, desc2_val, open2_val, sem2_val) if not is_absent2 else 0
            
            # Calculate final internal marks
            final_internal = calculate_final_internal_marks(mid1, mid2)

            # Check if record exists
            record = db.query(InternalMarks).filter(
                InternalMarks.sid == student.id,
                InternalMarks.subject_code == subject_code,
                InternalMarks.semester == semester
            ).first()

            if record:
                # Update existing
                record.objective1 = obj1_val
                record.descriptive1 = desc1_val
                record.openbook1 = open1_val
                record.seminar1 = sem1_val
                record.objective2 = obj2_val
                record.descriptive2 = desc2_val
                record.openbook2 = open2_val
                record.seminar2 = sem2_val
                record.mid1 = mid1
                record.mid2 = mid2
                record.final_internal_marks = final_internal
                record.entered_by = faculty_email
            else:
                # Create new
                db.add(InternalMarks(
                    sid=student.id,
                    srno=roll_no,
                    subject_code=subject_code,
                    subject_name=subject_name,
                    semester=semester,
                    objective1=obj1_val,
                    descriptive1=desc1_val,
                    openbook1=open1_val,
                    seminar1=sem1_val,
                    objective2=obj2_val,
                    descriptive2=desc2_val,
                    openbook2=open2_val,
                    seminar2=sem2_val,
                    mid1=mid1,
                    mid2=mid2,
                    final_internal_marks=final_internal,
                    entered_by=faculty_email
                ))
            
            processed_count += 1
            
        except Exception as e:
            print(f"Error processing row: {e}")
            continue

    db.commit()
    print(f"Uploaded and calculated internal marks for {processed_count} records")


def upload_external_marks_excel(
    db,
    file,
    batch: str,
    semester: int,
    admin_email: str,
    branch: str = None,
    section: str = None
):
    """
    Admin uploads external marks (out of 70).
    
    Excel columns: rollno, subjectcode, subjectname, credits, externalmarks
    
    Calculates:
    - Semester marks = internal (30) + external (70) = out of 100
    - Grade from semester marks
    - SGPA and CGPA
    """
    file.file.seek(0)
    wb = load_workbook(BytesIO(file.file.read()))
    sheet = wb.active

    processed_students = {}  # Track for SGPA/CGPA calculation
    processed_count = 0
    
    for roll_no, subject_code, subject_name, credits, external_marks in iter_external_marks_sheet_rows(sheet):
        try:
            if not roll_no or not subject_code:
                continue
                
            roll_no = str(roll_no).strip()
            subject_code = str(subject_code).strip()
            subject_name = str(subject_name).strip() if subject_name else subject_code
            
            try:
                credits = int(credits) if credits is not None and credits != "" else 0
            except (ValueError, TypeError):
                credits = 0
                
            try:
                external_val = int(float(external_marks)) if external_marks is not None and external_marks != "" else 0
            except (ValueError, TypeError):
                external_val = 0

            # Find student
            student = db.query(Student).filter(
                Student.roll_no == roll_no
            ).first()
            
            if not student:
                continue

            # Find academic record (strict filters first, then batch-only fallback)
            academic_query = db.query(Academic).filter(
                Academic.sid == student.id,
                Academic.batch == batch
            )
            
            if branch and branch != "All":
                academic_query = academic_query.filter(Academic.branch == branch)
            
            if section and section != "All":
                academic_query = academic_query.filter(Academic.section == section)
                
            academic = academic_query.first()
            if not academic:
                academic = (
                    db.query(Academic)
                    .filter(Academic.sid == student.id, Academic.batch == batch)
                    .first()
                )
            if not academic:
                continue

            student_branch = academic.branch
            student_section = academic.section

            # Get internal marks for this student-subject
            internal = db.query(InternalMarks).filter(
                InternalMarks.sid == student.id,
                InternalMarks.subject_code == subject_code,
                InternalMarks.semester == semester
            ).first()

            internal_marks_val = internal.final_internal_marks if internal else 0
            
            # Calculate semester marks (internal 30 + external 70)
            semester_marks = calculate_semester_marks(internal_marks_val, external_val)
            
            # Calculate grade
            grade_letter, grade_points = grade_from_percentage(semester_marks)

            # Store external marks
            record = db.query(ExternalMarks).filter(
                ExternalMarks.sid == student.id,
                ExternalMarks.subject_code == subject_code,
                ExternalMarks.semester == semester,
                ExternalMarks.batch == batch
            ).first()

            if record:
                record.subject_name = subject_name
                record.credits = credits
                record.external_marks = external_val
                record.grade = grade_letter
                record.entered_by = admin_email
            else:
                db.add(ExternalMarks(
                    sid=student.id,
                    srno=roll_no,
                    subject_code=subject_code,
                    subject_name=subject_name,
                    credits=credits,
                    external_marks=external_val,
                    batch=batch,
                    branch=student_branch,
                    section=student_section,
                    semester=semester,
                    grade=grade_letter,
                    entered_by=admin_email
                ))

            # Store in CourseGrade
            course_grade = db.query(CourseGrade).filter(
                CourseGrade.sid == student.id,
                CourseGrade.subject_code == subject_code,
                CourseGrade.semester == semester,
                CourseGrade.batch == batch
            ).first()

            if course_grade:
                # Update
                course_grade.internal_marks = internal_marks_val
                course_grade.external_marks = external_val
                course_grade.semester_marks = semester_marks
                course_grade.grade_letter = grade_letter
                course_grade.grade_points = grade_points
            else:
                # Create new
                db.add(CourseGrade(
                    sid=student.id,
                    srno=roll_no,
                    subject_code=subject_code,
                    subject_name=subject_name,
                    credits=credits,
                    internal_marks=internal_marks_val,
                    external_marks=external_val,
                    semester_marks=semester_marks,
                    grade_letter=grade_letter,
                    grade_points=grade_points,
                    batch=batch,
                    branch=student_branch,
                    section=student_section,
                    semester=semester
                ))
            
            processed_count += 1

            # Track for SGPA calculation
            student_key = student.id
            if student_key not in processed_students:
                processed_students[student_key] = {
                    "total_points": 0.0,
                    "total_credits": 0,
                    "has_fail": False,
                    "branch": student_branch,
                    "section": student_section,
                }
            
            processed_students[student_key]["total_points"] += grade_points * credits
            processed_students[student_key]["total_credits"] += credits
            
            if grade_letter == "F":
                processed_students[student_key]["has_fail"] = True
                
        except Exception as e:
            print(f"Error processing row: {e}")
            continue

    db.commit()

    # Calculate and store SGPA/CGPA
    print(f"Calculating SGPA/CGPA for {len(processed_students)} students...")
    
    for student_id, data in processed_students.items():
        # Calculate SGPA for this semester
        if data["total_credits"] > 0:
            sgpa = calc_sgpa(data["total_points"], data["total_credits"])
        else:
            sgpa = 0.0

        result_status = "FAIL" if data["has_fail"] else "PASS"

        # Get all previous SGPA values for CGPA calculation
        all_semester_grades = db.query(SemesterGrade).filter(
            SemesterGrade.sid == student_id,
            SemesterGrade.batch == batch
        ).order_by(SemesterGrade.semester).all()

        # Calculate CGPA including this semester
        total_cgpa_points = 0.0
        total_cgpa_credits = 0
        
        for sem in all_semester_grades:
            total_cgpa_points += sem.sgpa * sem.total_credits
            total_cgpa_credits += sem.total_credits
        
        # Add current semester
        total_cgpa_points += sgpa * data["total_credits"]
        total_cgpa_credits += data["total_credits"]
        
        cgpa = calc_sgpa(total_cgpa_points, total_cgpa_credits) if total_cgpa_credits > 0 else sgpa
        cgpa_percentage = cgpa_to_percentage(cgpa)

        # Store semester grade
        student_record = db.query(Student).filter(Student.id == student_id).first()
        
        sem_grade = db.query(SemesterGrade).filter(
            SemesterGrade.sid == student_id,
            SemesterGrade.semester == semester,
            SemesterGrade.batch == batch
        ).first()

        if sem_grade:
            # Update
            sem_grade.sgpa = sgpa
            sem_grade.cgpa = cgpa
            sem_grade.cgpa_percentage = cgpa_percentage
            sem_grade.total_credits = data["total_credits"]
            sem_grade.result_status = result_status
        else:
            # Create new
            db.add(SemesterGrade(
                sid=student_id,
                srno=student_record.roll_no if student_record else "",
                batch=batch,
                branch=data["branch"],
                section=data["section"],
                semester=semester,
                sgpa=sgpa,
                cgpa=cgpa,
                cgpa_percentage=cgpa_percentage,
                total_credits=data["total_credits"],
                result_status=result_status
            ))

    db.commit()
    print(f"✓ Uploaded external marks for {processed_count} records and calculated grades")
