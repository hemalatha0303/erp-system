
from app.models.academic import Academic
from app.models.hostel_room import HostelRoom
from app.models.hostel_allocation import HostelAllocation
from app.models.student import Student
from datetime import date
from app.models.payment import Payment
from app.models.fee_structure import FeeStructure
from openpyxl import load_workbook
from io import BytesIO
from sqlalchemy import func

def upload_hostel_rooms_excel(db, file):

    file.file.seek(0)
    wb = load_workbook(BytesIO(file.file.read()))
    sheet = wb.active

    for row in sheet.iter_rows(min_row=2, values_only=True):
        room_number, sharing, room_type, capacity = row

        if not room_number:
            continue

        
        exists = db.query(HostelRoom).filter(
            HostelRoom.room_number == str(room_number)
        ).first()

        if exists:
            continue

        room = HostelRoom(
            room_number=str(room_number),
            sharing=int(sharing),
            room_type=str(room_type).upper(),
            capacity=int(capacity),
            occupied=0,
            is_active=True
        )
        db.add(room)

    db.commit()

def allocate_student_hostel(db, req, admin_email):

    student = db.query(Student).filter(
        Student.roll_no == req.roll_no
    ).first()

    if student.residence_type != "HOSTELER":
        raise Exception("Student is not a hosteler")

    room = db.query(HostelRoom).filter(
        HostelRoom.room_number == req.room_number,
        HostelRoom.is_active == True
    ).first()

    if room.occupied >= room.capacity:
        raise Exception("Room is full")

    existing = db.query(HostelAllocation).filter(
        HostelAllocation.student_id == student.id,
        HostelAllocation.status == "ALLOCATED"
    ).first()

    if existing:
        raise Exception("Student already allocated")

    allocation = HostelAllocation(
        student_id=student.id,
        room_id=room.id,
        allocated_date=date.today(),
        status="ALLOCATED",
        allocated_by=admin_email
    )

    room.occupied += 1

    db.add(allocation)
    db.commit()
def get_student_hostel_details(db, student_email):

    student = db.query(Student).filter(
        Student.user_email == student_email
    ).first()
    academic = db.query(Academic).filter(
        Academic.sid == student.id
    ).first()
    allocation = (
        db.query(HostelAllocation, HostelRoom)
        .join(HostelRoom, HostelRoom.id == HostelAllocation.room_id)
        .filter(
            HostelAllocation.student_id == student.id,
            HostelAllocation.status == "ALLOCATED"
        )
        .first()
    )

    fee_struct = db.query(FeeStructure).filter(
        FeeStructure.quota == academic.quota,
        FeeStructure.residence_type == "HOSTELER",
        FeeStructure.year == academic.year
    ).first()
    payment = db.query(Payment).filter(
        Payment.srno == student.roll_no,  
        Payment.year == academic.year,
        Payment.fee_type == "HOSTEL"
    ).first()
    if not allocation:
        return {"hostel": None}
    if not fee_struct :
        return {"hostel": None}
    if payment.amount_paid is None:
        payment.amount_paid = 0
    alloc, room = allocation

    return {
        "room_number": room.room_number,
        "room_type": room.room_type,
        "sharing": room.sharing,
        "allocated_date": alloc.allocated_date,
        "fee": {
            "total": float(fee_struct.hostel_fee),
            "paid": float(payment.amount_paid),
            "due": float(fee_struct.hostel_fee - payment.amount_paid),
            "status": payment.status,
            "last_paid_date": payment.payment_date
        }
    }

def vacate_hostel_room(db, req, admin_email):

    student = db.query(Student).filter(
        Student.roll_no == req.roll_no
    ).first()
    allocation = db.query(HostelAllocation).filter(
        HostelAllocation.student_id == student.id,
        HostelAllocation.status == "ALLOCATED"
    ).first()
    if not allocation:
        raise Exception("No active allocation found for the student")
    
    room = db.query(HostelRoom).filter(
        HostelRoom.room_number == req.room_number
    ).first()
    allocation.status = "VACATED"
    allocation.vacated_date = date.today()
    room.occupied -= 1
    db.commit()

# ==================== NEW CRUD OPERATIONS ====================

def get_hostel_statistics(db):
    """Get hostel statistics: total rooms, allocated, vacated, not allocated"""
    total_rooms = db.query(HostelRoom).filter(HostelRoom.is_active == True).count()
    total_allocated = db.query(HostelAllocation).filter(HostelAllocation.status == "ALLOCATED").count()
    total_vacated = db.query(HostelAllocation).filter(HostelAllocation.status == "VACATED").count()
    
    # Not allocated = total rooms - currently occupied rooms
    occupied_rooms = db.query(HostelRoom).filter(HostelRoom.is_active == True).with_entities(
        func.sum(HostelRoom.occupied)
    ).scalar() or 0
    
    return {
        "total_rooms": total_rooms,
        "allocated": total_allocated,
        "vacated": total_vacated,
        "not_allocated": total_rooms - occupied_rooms
    }

def get_all_hostel_rooms(db):
    """Get all hostel rooms with their details"""
    from sqlalchemy import func
    rooms = db.query(HostelRoom).filter(HostelRoom.is_active == True).all()
    
    rooms_data = []
    for room in rooms:
        room_data = {
            "id": room.id,
            "room_number": room.room_number,
            "sharing": room.sharing,
            "room_type": room.room_type,
            "capacity": room.capacity,
            "occupied": room.occupied,
            "available": room.capacity - room.occupied
        }
        rooms_data.append(room_data)
    
    return rooms_data

def get_all_hostel_allocations(db):
    """Get all hostel allocations with student and room details"""
    allocations = db.query(
        HostelAllocation,
        Student,
        Academic,
        HostelRoom
    ).join(
        Student, Student.id == HostelAllocation.student_id
    ).join(
        Academic, Academic.sid == Student.id
    ).join(
        HostelRoom, HostelRoom.id == HostelAllocation.room_id
    ).all()
    
    allocations_data = []
    for alloc, student, academic, room in allocations:
        alloc_data = {
            "allocation_id": alloc.id,
            "student_id": student.id,
            "roll_no": student.roll_no,
            "name": f"{student.first_name} {student.last_name}",
            "year": academic.year,
            "branch": academic.branch,
            "section": academic.section,
            "room_number": room.room_number,
            "room_type": room.room_type,
            "sharing": room.sharing,
            "status": alloc.status,
            "allocated_date": str(alloc.allocated_date) if alloc.allocated_date else None,
            "vacated_date": str(alloc.vacated_date) if alloc.vacated_date else None,
            "allocated_by": alloc.allocated_by
        }
        allocations_data.append(alloc_data)
    
    return allocations_data

def create_hostel_room(db, room_number: str, sharing: int, room_type: str, capacity: int):
    """Create a new hostel room"""
    exists = db.query(HostelRoom).filter(HostelRoom.room_number == room_number).first()
    if exists:
        raise Exception(f"Room {room_number} already exists")
    
    room = HostelRoom(
        room_number=room_number,
        sharing=sharing,
        room_type=room_type.upper(),
        capacity=capacity,
        occupied=0,
        is_active=True
    )
    db.add(room)
    db.commit()
    db.refresh(room)
    return room

def update_hostel_room(db, room_id: int, sharing: int = None, room_type: str = None, capacity: int = None):
    """Update hostel room details"""
    room = db.query(HostelRoom).filter(HostelRoom.id == room_id).first()
    if not room:
        raise Exception("Room not found")
    
    if sharing is not None:
        room.sharing = sharing
    if room_type is not None:
        room.room_type = room_type.upper()
    if capacity is not None:
        room.capacity = capacity
    
    db.commit()
    db.refresh(room)
    return room

def delete_hostel_room(db, room_id: int):
    """Soft delete a hostel room"""
    room = db.query(HostelRoom).filter(HostelRoom.id == room_id).first()
    if not room:
        raise Exception("Room not found")
    
    # Check if room has any active allocations
    active_alloc = db.query(HostelAllocation).filter(
        HostelAllocation.room_id == room_id,
        HostelAllocation.status == "ALLOCATED"
    ).first()
    
    if active_alloc:
        raise Exception("Cannot delete room with active allocations")
    
    room.is_active = False
    db.commit()

def update_allocation_status(db, allocation_id: int, new_status: str, admin_email: str):
    """Update allocation status (ALLOCATED/VACATED/NOT_ALLOCATED)"""
    allocation = db.query(HostelAllocation).filter(HostelAllocation.id == allocation_id).first()
    if not allocation:
        raise Exception("Allocation not found")
    
    room = db.query(HostelRoom).filter(HostelRoom.id == allocation.room_id).first()
    
    old_status = allocation.status
    allocation.status = new_status
    
    # Update room occupancy
    if old_status == "ALLOCATED" and new_status == "VACATED":
        allocation.vacated_date = date.today()
        room.occupied -= 1
    elif old_status == "VACATED" and new_status == "ALLOCATED":
        allocation.vacated_date = None
        room.occupied += 1
    
    db.commit()
    db.refresh(allocation)
    return allocation

def allocate_student_to_hostel_ui(db, student_id: int, room_id: int, admin_email: str):
    """Allocate student to hostel room from UI"""
    student = db.query(Student).filter(Student.id == student_id).first()
    if not student:
        raise Exception("Student not found")
    
    if student.residence_type != "HOSTELER":
        raise Exception("Student is not a hosteler")
    
    room = db.query(HostelRoom).filter(HostelRoom.id == room_id, HostelRoom.is_active == True).first()
    if not room:
        raise Exception("Room not found")
    
    if room.occupied >= room.capacity:
        raise Exception("Room is full")
    
    # Check for existing allocation
    existing = db.query(HostelAllocation).filter(
        HostelAllocation.student_id == student_id,
        HostelAllocation.status == "ALLOCATED"
    ).first()
    
    if existing:
        raise Exception("Student already allocated")
    
    allocation = HostelAllocation(
        student_id=student_id,
        room_id=room_id,
        allocated_date=date.today(),
        status="ALLOCATED",
        allocated_by=admin_email
    )
    
    room.occupied += 1
    db.add(allocation)
    db.commit()
    db.refresh(allocation)
    return allocation

