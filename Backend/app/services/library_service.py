from openpyxl import load_workbook
from io import BytesIO
from app.models.academic import Academic
from app.models.library_books import LibraryBooks
from app.models.library_issue import LibraryIssue
from app.models.student import Student
from sqlalchemy.orm import Session
def get_student_library_books(db, student_email: str, semester: int):

    student = db.query(Student).filter(
        Student.user_email == student_email
    ).first()

    if not student:
        raise Exception("Student not found")

    records = (
        db.query(LibraryIssue, LibraryBooks)
        .join(LibraryBooks, LibraryBooks.code == LibraryIssue.book_code)
        .filter(
            LibraryIssue.srno == student.roll_no,
            LibraryIssue.semester == semester
        )
        .order_by(LibraryIssue.issued_date.desc())
        .all()
    )

    result = []
    for issue, book in records:
        result.append({
            "book_code": book.code,
            "title": book.title,
            "author": book.author,
            "issued_date": issue.issued_date,
            "return_date": issue.return_date,
            "status": issue.status
        })

    return {
        "semester": semester,
        "books": result
    }




def issue_books_to_student(db, req):

    student = db.query(Student).filter(
        Student.roll_no == req.srno
    ).first()
    if not student:
        raise Exception("Student not found")
    academic = db.query(Academic).filter(
        Academic.srno == req.srno   
    ).order_by(Academic.year.desc()).first()
    for book_code in req.book_codes:
        book = db.query(LibraryBooks).filter(
            LibraryBooks.code == book_code
        ).first()
        if not book:
            raise Exception(f"Book {book_code} not found")

        if book.available_copies <= 0:
            raise Exception(f"Book {book_code} not available")

        issue = LibraryIssue(
            srno=req.srno,
            book_code=book.code,
            
            semester=req.semester,
            year=academic.year,
            status="ISSUED",
            updated_by="Library Admin",
            issued_date=req.issued_date,
            expected_return_date=req.expected_return_date

        )

        book.available_copies -= 1
        db.add(issue)

    db.commit()


def upload_books_excel(db, file):
    file.file.seek(0)
    wb = load_workbook(BytesIO(file.file.read()))
    sheet = wb.active

    for row in sheet.iter_rows(min_row=2, values_only=True):
        code, title, author, copies = row

        book = db.query(LibraryBooks).filter(
            LibraryBooks.code == code
        ).first()

        if book:
            book.available_copies += int(copies)
        else:
            db.add(LibraryBooks(
                code=code,
                title=title,
                author=author,
                available_copies=copies
            ))

    db.commit()


def assign_book(db, req, admin_email):

    book = db.query(LibraryBooks).filter(
        LibraryBooks.code == req.book_code
    ).first()

    if not book or book.available_copies <= 0:
        raise Exception("Book not available")

    db.add(LibraryIssue(
        srno=req.srno,
        book_code=req.book_code,
        semester=req.semester,
        year=req.year,
        status="ISSUED",
        updated_by=admin_email
    ))

    book.available_copies -= 1
    db.commit()

def return_books(db: Session, req, admin_email: str):

    issues = db.query(LibraryIssue).filter(
        LibraryIssue.srno == req.srno,
        LibraryIssue.semester == req.semester,
        
        LibraryIssue.status == "ISSUED",
        LibraryIssue.book_code.in_(req.book_codes)
    ).all()

    if not issues:
        raise Exception("No issued books found")

    for issue in issues:
        
        issue.status = "RETURNED"
        issue.return_date = req.return_date
        issue.updated_by = admin_email

        
        book = db.query(LibraryBooks).filter(
            LibraryBooks.code == issue.book_code
        ).first()

        if book:
            book.available_copies += 1

    db.commit()

    return {
        "message": "Books returned successfully",
        "returned_count": len(issues)
    }

def get_pending_books(db, srno: str, semester: int):

    return db.query(LibraryIssue).filter(
        LibraryIssue.srno == srno,
        LibraryIssue.semester == semester,
        LibraryIssue.status == "ISSUED"
    ).all()
