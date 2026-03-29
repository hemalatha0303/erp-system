from openpyxl import load_workbook
from io import BytesIO
from datetime import date
from sqlalchemy.exc import IntegrityError
from app.models.student import Student
from app.models.academic import Academic
from app.models.payment import Payment
from app.models.fee_structure import FeeStructure
from app.models.faculty import Faculty
from app.models.hod import HODProfile
from app.models.user import User
from app.core.security import hash_password
from app.services.bulk_user_service import generate_password

def _ensure_user(db, email, role, personal_email=None):
    existing_user = db.query(User).filter(User.email == email).first()
    if existing_user:
        if personal_email and not existing_user.personal_email:
            existing_user.personal_email = personal_email
        return existing_user
    password = generate_password()
    user = User(
        email=email,
        password=hash_password(password),
        role=role,
        personal_email=personal_email,
        is_active=True
    )
    db.add(user)
    return user

def _apply_if_value(model, attr, value):
    if value is not None and value != "":
        setattr(model, attr, value)

def process_excel(file, db, admin_email, role="STUDENT", selected_semester=1):
    """
    Process Excel file for bulk data upload.
    
    Args:
        file: Uploaded file
        db: Database session
        admin_email: Admin's email
        role: Type of user (STUDENT, FACULTY, HOD)
        selected_semester: Semester for students (default 1)
    """
    try:
        file_bytes = BytesIO(file.file.read())
        wb = load_workbook(file_bytes)
        sheet = wb.active

        if role == "STUDENT":
            _process_student_excel(sheet, db, admin_email, selected_semester)
        elif role == "FACULTY":
            _process_faculty_excel(sheet, db)
        elif role == "HOD":
            _process_hod_excel(sheet, db)
        
        db.commit()
    except IntegrityError as e:
        db.rollback()
        error_msg = str(e).lower()
        if "duplicate" in error_msg:
            raise ValueError(f"Duplicate entry found: {str(e)[:100]}")
        raise
    except Exception as e:
        db.rollback()
        raise


def _process_student_excel(sheet, db, admin_email, selected_semester):
    """Process student data from Excel."""
    for r in sheet.iter_rows(min_row=2, values_only=True):
        # Expected columns: roll_no, email, first_name, last_name, personal_mail, gender, 
        # residence_type, mobile_no, parent_mobile_no, branch, section, batch, course, 
        # quota, admission_date, address, parent_name
        
        if not r[1]:  # Skip if email is empty
            continue
            
        (
            roll, email, fn, ln, personal_email, gender, residence, mob, pmob,
            branch, section, batch, course, quota, adm, address, parent_name
        ) = r[:17]
        
        if not email:
            continue
        
        email = str(email).strip().lower()
        
        # Check if student already exists
        existing = db.query(Student).filter(Student.user_email == email).first()
        if existing:
            continue
        roll_no_str = str(roll).strip() if roll is not None else None
        if roll_no_str:
            existing_roll = db.query(Student).filter(Student.roll_no == roll_no_str).first()
            if existing_roll:
                continue

        student = Student(
            roll_no=roll_no_str,
            user_email=email,
            personal_email=personal_email,
            first_name=fn,
            last_name=ln,
            gender=gender,
            residence_type=residence,
            mobile_no=mob,
            parent_mobile_no=pmob,
            address=address,
            parentname=parent_name,
            branch=branch,
            section=section,
            batch=batch,
            course=course,
            quota=quota,
            admission_date=adm
        )
        db.add(student)
        db.flush()

        # Create Academic record for the selected semester
        academic = Academic(
            sid=student.id,
            srno=roll,
            user_email=email,
            branch=branch,
            batch=batch,
            course=course,
            year=_calculate_year_from_batch(batch),
            semester=selected_semester,
            section=section,
            quota=quota,
            admission_date=adm,
            status="ACTIVE"
        )
        db.add(academic)

        # Create payment records
        _create_payment_records(db, roll, email, selected_semester, residence, quota, admin_email)


def _process_faculty_excel(sheet, db):
    """Process faculty data from Excel."""
    # Get header row
    header = [cell.value for cell in sheet[1]]
    seen_emails = set()
    
    # Map column indices from header
    try:
        email_col = header.index("email")
        first_name_col = header.index("first_name") if "first_name" in header else None
        last_name_col = header.index("last_name") if "last_name" in header else None
        mobile_no_col = header.index("mobile_no") if "mobile_no" in header else None
        personal_email_col = header.index("personal_email") if "personal_email" in header else None
        address_col = header.index("address") if "address" in header else None
        qualification_col = header.index("qualification") if "qualification" in header else None
        years_exp_col = header.index("years_of_experience") if "years_of_experience" in header else None
        subject_code_col = header.index("subject_code") if "subject_code" in header else None
        subject_name_col = header.index("subject_name") if "subject_name" in header else None
        branch_col = header.index("branch") if "branch" in header else None
    except ValueError as e:
        raise ValueError(f"Excel missing required column: {str(e)}")

    for row in sheet.iter_rows(min_row=2):
        email = row[email_col].value
        
        if not email:
            continue
        
        email = str(email).strip().lower()
        if email in seen_emails:
            continue
        seen_emails.add(email)

        # Extract all fields with proper type conversion
        first_name = str(row[first_name_col].value).strip() if first_name_col is not None and row[first_name_col].value else None
        last_name = str(row[last_name_col].value).strip() if last_name_col is not None and row[last_name_col].value else None
        mobile_no = str(row[mobile_no_col].value).strip() if mobile_no_col is not None and row[mobile_no_col].value else None
        personal_email = str(row[personal_email_col].value).strip() if personal_email_col is not None and row[personal_email_col].value else None
        address = str(row[address_col].value).strip() if address_col is not None and row[address_col].value else None
        qualification = str(row[qualification_col].value).strip() if qualification_col is not None and row[qualification_col].value else None
        
        years_exp = None
        if years_exp_col is not None and row[years_exp_col].value:
            try:
                years_exp = int(row[years_exp_col].value) if isinstance(row[years_exp_col].value, (int, float)) else None
            except (ValueError, TypeError):
                years_exp = None
        
        subject_code = str(row[subject_code_col].value).strip() if subject_code_col is not None and row[subject_code_col].value else None
        subject_name = str(row[subject_name_col].value).strip() if subject_name_col is not None and row[subject_name_col].value else None
        branch = str(row[branch_col].value).strip() if branch_col is not None and row[branch_col].value else None

        _ensure_user(db, email, "FACULTY", personal_email)

        # Check if faculty already exists
        existing = db.query(Faculty).filter(Faculty.user_email == email).first()
        if existing:
            _apply_if_value(existing, "personal_email", personal_email)
            _apply_if_value(existing, "first_name", first_name)
            _apply_if_value(existing, "last_name", last_name)
            _apply_if_value(existing, "mobile_no", mobile_no)
            _apply_if_value(existing, "address", address)
            _apply_if_value(existing, "qualification", qualification)
            _apply_if_value(existing, "experience", years_exp)
            _apply_if_value(existing, "subject_code", subject_code)
            _apply_if_value(existing, "subject_name", subject_name)
            _apply_if_value(existing, "branch", branch)
            continue

        faculty = Faculty(
            user_email=email,
            personal_email=personal_email,
            first_name=first_name,
            last_name=last_name,
            mobile_no=mobile_no,
            address=address,
            qualification=qualification,
            experience=years_exp,
            subject_code=subject_code,
            subject_name=subject_name,
            branch=branch
        )
        db.add(faculty)


def _process_hod_excel(sheet, db):
    """Process HOD data from Excel."""
    # Get header row
    header = [cell.value for cell in sheet[1]]
    seen_emails = set()
    
    # Map column indices from header
    try:
        email_col = header.index("email")
        first_name_col = header.index("first_name") if "first_name" in header else None
        last_name_col = header.index("last_name") if "last_name" in header else None
        mobile_no_col = header.index("mobile_no") if "mobile_no" in header else None
        personal_email_col = header.index("personal_email") if "personal_email" in header else None
        address_col = header.index("address") if "address" in header else None
        qualification_col = header.index("qualification") if "qualification" in header else None
        years_exp_col = header.index("years_of_experience") if "years_of_experience" in header else None
        branch_col = header.index("branch") if "branch" in header else None
    except ValueError as e:
        raise ValueError(f"Excel missing required column: {str(e)}")

    for row in sheet.iter_rows(min_row=2):
        email = row[email_col].value
        
        if not email:
            continue
        
        email = str(email).strip().lower()
        if email in seen_emails:
            continue
        seen_emails.add(email)

        # Extract all fields with proper type conversion
        first_name = str(row[first_name_col].value).strip() if first_name_col is not None and row[first_name_col].value else None
        last_name = str(row[last_name_col].value).strip() if last_name_col is not None and row[last_name_col].value else None
        mobile_no = str(row[mobile_no_col].value).strip() if mobile_no_col is not None and row[mobile_no_col].value else None
        personal_email = str(row[personal_email_col].value).strip() if personal_email_col is not None and row[personal_email_col].value else None
        address = str(row[address_col].value).strip() if address_col is not None and row[address_col].value else None
        qualification = str(row[qualification_col].value).strip() if qualification_col is not None and row[qualification_col].value else None
        
        years_exp = None
        if years_exp_col is not None and row[years_exp_col].value:
            try:
                years_exp = int(row[years_exp_col].value) if isinstance(row[years_exp_col].value, (int, float)) else None
            except (ValueError, TypeError):
                years_exp = None
        
        branch = str(row[branch_col].value).strip() if branch_col is not None and row[branch_col].value else None
        
        _ensure_user(db, email, "HOD", personal_email)

        # Check if HOD already exists
        existing = db.query(HODProfile).filter(HODProfile.email == email).first()
        if existing:
            _apply_if_value(existing, "personal_email", personal_email)
            _apply_if_value(existing, "first_name", first_name)
            _apply_if_value(existing, "last_name", last_name)
            _apply_if_value(existing, "mobile_no", mobile_no)
            _apply_if_value(existing, "address", address)
            _apply_if_value(existing, "qualification", qualification)
            _apply_if_value(existing, "experience", years_exp)
            _apply_if_value(existing, "branch", branch)
            continue

        hod = HODProfile(
            email=email,
            personal_email=personal_email,
            first_name=first_name,
            last_name=last_name,
            mobile_no=mobile_no,
            address=address,
            qualification=qualification,
            experience=years_exp,
            branch=branch
        )
        db.add(hod)


def _calculate_year_from_batch(batch):
    """Calculate year from batch (e.g., 2022-26 batch -> year 1)."""
    try:
        if batch and '-' in str(batch):
            start_year = int(str(batch).split('-')[0])
            current_year = date.today().year
            year = min(current_year - start_year + 1, 4)
            return max(year, 1)
    except:
        pass
    return 1


def _create_payment_records(db, roll, email, semester, residence, quota, admin_email):
    """Create initial payment records for student."""
    year = 1  # Default year, should be calculated from batch if needed
    
    fee = db.query(FeeStructure).filter(
        FeeStructure.quota == quota,
        FeeStructure.residence_type == residence,
        FeeStructure.year == year,
    ).first()

    if not fee:
        # Skip if fee structure not found
        return

    if fee.tuition_fee > 0:
        db.add(Payment(
            receipt_id=f"INIT-T-{roll}",
            srno=roll,
            student_email=email,
            fee_type="TUITION",
            amount_paid=0,
            year=year,
            semester=semester,
            payment_date=None,
            status="PENDING",
            updated_by=admin_email
        ))

    if residence == "DAY_SCHOLAR":
        db.add(Payment(
            receipt_id=f"INIT-B-{roll}",
            srno=roll,
            student_email=email,
            fee_type="BUS",
            amount_paid=0,
            year=year,
            semester=semester,
            payment_date=date.today(),
            status="PENDING",
            updated_by=admin_email
        ))

    if residence == "HOSTELER":
        db.add(Payment(
            receipt_id=f"INIT-H-{roll}",
            srno=roll,
            student_email=email,
            fee_type="HOSTEL",
            amount_paid=0,
            year=year,
            semester=semester,
            payment_date=date.today(),
            status="PENDING",
            updated_by=admin_email
        ))
